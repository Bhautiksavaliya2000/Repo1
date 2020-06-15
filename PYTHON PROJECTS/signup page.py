import tkinter as tk
from tkinter import *
from tkinter.messagebox import *
from tkinter import ttk
from openpyxl import *
import openpyxl
from openpyxl import load_workbook
import hashlib



def getvalue():
    if (username_var.get() == ''):
        print(showinfo("warning", "* fields must be filled."))
    elif pass_var.get() == '':
        print(showinfo("warning", "* fields must be filled."))
    elif pass_var.get() != repass_var.get():
        print(showinfo("warning", "both password must be same."))

    #get every entry's value.
    else:
        name = username_var.get()
        mail = email_var.get()
        pas = pass_var.get()
        pas2 = repass_var.get()
        gender = gender_var.get()
        user_type = usertype.get()

        # to load workbook 

        wb = load_workbook(r'G:\PYTHON\GUI\signup page\user_details.xlsx')
        ws = wb.active            # to get worksheet
        ws_copy = ws
        ws['A1'] = 'USERNAME'
        ws['B1'] = 'EMAIL ADDRESS'
        ws['C1'] = 'PASSWORD'
        ws['D1'] = 'GENDER'
        ws['E1'] = 'USER TYPE'
        
        hash_encode = hashlib.md5(pas.encode())
        hash_decode = hash_encode.hexdigest()
        attrib_list = [name,mail,hash_decode,gender,user_type]
        ws.append(attrib_list)

        wb.save(filename=r'G:\PYTHON\GUI\signup page\user_details.xlsx')
        
        username.delete(0,tk.END)
        email.delete(0,tk.END)
        password.delete(0,tk.END)
        repass.delete(0,tk.END)

def switchstate():
    if check_btn.get() == 1:
       signup_button.config(state = tk.ACTIVE) 
    elif check_btn.get() == 0:
       signup_button.config(state = tk.DISABLED)             

# defining new window inside button command.
def signin_win():
    
    signin_window = tk.Tk()
    signin_window.title("Signin")
    signin_window.geometry("310x250")
    signin_window.maxsize(300,300)
    signin_window.configure(bg = '#F5F5DC')
    
    def search_details():
        u_name = username_entry.get()
        p_word = password_entry.get()

        wb_copy = load_workbook(r'G:\PYTHON\GUI\signup page\user_details.xlsx')
        ws_copy = wb_copy.active

        success_frame = tk.Frame(signin_window,height=20,bg='#F5F5DC',)
        success_frame.pack(side=TOP)

        success_label = tk.Label(success_frame,text="",fg='red')
        success_label.pack(side=BOTTOM,anchor=SW)
        
        def search_for_pass(no):            # If username is matched then only it go for search password to that perticular row.
            pass_obj = ws_copy.cell(row=no,column=3)
            encode_pass = hashlib.md5(p_word.encode())
            hex_pass = encode_pass.hexdigest()
            if hex_pass == pass_obj.value:
                success_label.configure(text='Login successful')                     
                username_entry.delete(0,tk.END)
                password_entry.delete(0,tk.END)               
            else:
                success_label.configure(text='incorrect password')                
                
        for i in range(1,ws_copy.max_row):
            name_obj = ws_copy.cell(row=i,column=1)
            if u_name == name_obj.value:
                search_for_pass(i)
                break
        else:
                success_label.configure(text='username not found')                
    
    def see_password():
        global count
        count+=1
        if count%2 == 0:
            password_entry.configure(show = '')
            show_button.configure(text='hide')
        else:
            password_entry.configure(show = '*')
            show_button.configure(text='see')


    logoframe = tk.Frame(signin_window,bg='red')
    logoframe.pack(side=TOP)
    

     # title label.
    logolabel = tk.Label(logoframe,text='Login',font='comicsansMS 35 bold',bg='#F5F5DC' ) 
    logolabel.pack()

    #creating frame for widgets.
    wintop_frame = tk.Frame(signin_window,bg='#F5F5DC',width=150,height=100)
    wintop_frame.pack(side=TOP,pady=15)
    #defining label
    username_label = ttk.Label(wintop_frame,text='Username :',background='#F5F5DC',foreground='#8B4513',font='comicsansMS 10 bold')
    username_label.pack(side=LEFT,anchor=N)
    #defining entry
    user_var = tk.StringVar()
    username_entry = tk.Entry(wintop_frame)
    username_entry.pack(side=LEFT)
    username_entry.focus()
    
    #defining frame.
    winbottom_frame = tk.Frame(signin_window)
    winbottom_frame.pack(pady=10)
    #defining label
    password_label = ttk.Label(winbottom_frame,text='Password :',background='#F5F5DC',foreground='#8B4513',font='comicsansMS 10 bold')
    password_label.pack(side=LEFT,anchor=N)
    #defining entry
    password_entry = tk.Entry(winbottom_frame,show='*',width=15)
    password_entry.pack(side=LEFT)
    #defining show button
    show_button = tk.Button(winbottom_frame,text='see',width=3,command=see_password)
    show_button.pack(side=LEFT)

    #defining button
    login_button_frame = tk.Frame(signin_window,bg='#F5F5DC')
    login_button_frame.pack(side=TOP)
    login_button = ttk.Button(login_button_frame,text='Signin',command=search_details)
    login_button.pack(side=BOTTOM,pady=20)

    signin_window.mainloop()


root = tk.Tk()
root.title("signup menu")
root.geometry("400x300")
root.config(bg='#F5F5DC')
                         
count = 0


#creating labels                                 #sticky = to stick label left side.
label1 = ttk.Label(root,text='Enter username *:',background='#F5F5DC',foreground='#8B4513',font='comicsansMS 10 bold')
label2 = ttk.Label(root,text='Enter email *:',background='#F5F5DC',foreground='#8B4513',font='comicsansMS 10 bold')
label3 = ttk.Label(root,text='Enter password *:',background='#F5F5DC',foreground='#8B4513',font='comicsansMS 10 bold')
label4 = ttk.Label(root,text='Confirm password *:',background='#F5F5DC',foreground='#8B4513',font='comicsansMS 10 bold')
label5 = ttk.Label(root,text='Choose your gender *:',background='#F5F5DC',foreground='#8B4513',font='comicsansMS 10 bold')
label6 = ttk.Label(root,text='Already registered?',background='#F5F5DC',foreground='#8B4513',font='comicsansMS 10 bold')
#creating entry fields
username_var = tk.StringVar()
username = ttk.Entry(root,textvariable=username_var)
username.focus()

email_var = tk.StringVar()
email = ttk.Entry(root,textvariable=email_var)

pass_var = tk.StringVar()
password = ttk.Entry(root,textvariable=pass_var,show='*')

repass_var = tk.StringVar()
repass = ttk.Entry(root,textvariable=repass_var,show='*')

#creating combobox
gender_var = tk.StringVar()
gender_combobox = ttk.Combobox(root,textvariable=gender_var,width=17,state='readonly',background='#F5F5DC',foreground='#8B4513')
gender_combobox['values'] = ('Male','Female','Others')
gender_combobox.current(0)

#creating radiobutton
usertype = tk.StringVar()                 #variable should be same to allow only one radiobutton select at a time.
radiobtn1 = ttk.Radiobutton(root,text='student',value='student',variable=usertype)
radiobtn2 = ttk.Radiobutton(root,text='graduate',value='graduate',variable=usertype)
radiobtn3 = ttk.Radiobutton(root,text='professional',value='professional',variable=usertype)

#creating cheakbox
check_btn = tk.IntVar()
checkbtn = ttk.Checkbutton(root,text='accept terms and conditions',variable=check_btn,command=switchstate)

#button for submit details.
signup_button = ttk.Button(root,text='sign up',command=getvalue,state=tk.DISABLED)

#button for open signin window.
signin_button = ttk.Button(root,text='signin',command=signin_win)
#setting position of each widgit.
label1.grid(row=0,column=0,sticky=tk.W,padx=10)
username.grid(row=0,column=1,padx=3,pady=5)

label2.grid(row=1,column=0,sticky=tk.W,padx=10)
email.grid(row=1,column=1,padx=3,pady=5)

label3.grid(row=2,column=0,sticky=tk.W,padx=10)
password.grid(row=2,column=1,padx=3,pady=5)

label4.grid(row=3,column=0,sticky=tk.W,padx=10)
repass.grid(row=3,column=1,padx=3,pady=5)

label5.grid(row=4,column=0,sticky=tk.W,padx=10)
gender_combobox.grid(row=4,column=1,padx=3,pady=3)

radiobtn1.grid(row=5,column=0,padx=3,pady=5)
radiobtn2.grid(row=5,column=1,padx=3,pady=5)
radiobtn3.grid(row=5,column=2,padx=3,pady=5)

checkbtn.grid(row=6,columnspan=3,sticky=tk.W,padx=3,pady=6)              # to merge three columns

signup_button.grid(row=7,column=1,padx=3,pady=5)
label6.grid(row=8,column=0,padx=3,pady=15)
signin_button.grid(row=8,column=1,pady=15)
root.mainloop()
